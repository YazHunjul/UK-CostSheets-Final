from openpyxl import load_workbook
import streamlit as st
import io
import os
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from copy import copy
import math
from openpyxl.worksheet.datavalidation import DataValidation

def generate_sheet(kitchen_info, genInfo):
    """
    Generate an Excel sheet based on kitchen information and return it as a BytesIO object.
    """
    try:
        # Load the workbook
        excel_path = '/Users/yazan/Desktop/Efficiency/UK-CostSheets-Final/app/costSheetGen/costSheetResources/Halton Cost Sheet Jan 2025.xlsx'
        wb = load_workbook(excel_path, data_only=False)  # Keep formulas
        
        # Get all CANOPY sheets
        canopy_sheets = [sheet for sheet in wb.sheetnames if 'CANOPY' in sheet]
        sheet_count = 0
        
        # Track totals for all sheets
        total_cost = 0
        total_selling_price = 0
        
        # Track sheet numbers for each kitchen
        kitchen_sheet_counts = {}
        
        for kitchen in kitchen_info:
            kitchen_name = kitchen['kitchen_name']
            if kitchen_name not in kitchen_sheet_counts:
                kitchen_sheet_counts[kitchen_name] = 1
            
            for floor in kitchen.get('floors', []):
                # Get the appropriate CANOPY sheet
                if sheet_count >= len(canopy_sheets):
                    st.error(f"Not enough CANOPY sheets in template for {kitchen['kitchen_name']}-{floor['floor_name']}")
                    return None
                
                current_sheet = wb[canopy_sheets[sheet_count]]
                
                # Store full names in hidden cells
                current_sheet['Z1'] = kitchen['kitchen_name']  # Full Level name
                current_sheet['Z2'] = floor['floor_name']    # Full Area name
                
                # Write full name to B1
                current_sheet['B1'] = f"{kitchen['kitchen_name']} - {floor['floor_name']}"
                
                # Format sheet name with counter
                sheet_title = f"CANOPY - {kitchen_name} ({kitchen_sheet_counts[kitchen_name]})"
                if len(sheet_title) > 31:
                    sheet_title = sheet_title[:27] + f" ({kitchen_sheet_counts[kitchen_name]})"  # Leave room for counter
                current_sheet.title = sheet_title
                
                kitchen_sheet_counts[kitchen_name] += 1
                
                # Hide the columns containing the full names
                current_sheet.column_dimensions['Z'].hidden = True
                
                # Write general info
                current_sheet['C3'] = genInfo.get('projectNum', '').title()
                current_sheet['C5'] = genInfo.get('customer', '').title()
                current_sheet['C7'] = genInfo.get('combined_initials', '')
                current_sheet['G3'] = genInfo.get('projectName', '').title()
                current_sheet['G5'] = genInfo.get('location', '').title()
                current_sheet['G7'] = genInfo.get('date', '')
                
                # Fill canopy data
                current_row = 12
                
                for canopy in floor.get('canopies', []):
                    # Get the reference number (try both possible keys)
                    ref_num = canopy.get('itemNum') or canopy.get('item_number') or canopy.get('reference_number', '')
                    
                    # Write reference number ONLY to B column
                    current_sheet[f'B{current_row}'] = ref_num
                    
                    # Write to main CANOPY sheet
                    current_sheet[f'C{current_row + 2}'] = canopy.get('configuration', '')
                    current_sheet[f'D{current_row + 2}'] = canopy.get('model', '')
                    current_sheet[f'E{current_row + 2}'] = canopy.get('width', '')
                    current_sheet[f'F{current_row + 2}'] = canopy.get('length', '')
                    current_sheet[f'G{current_row + 2}'] = canopy.get('height', '')
                    current_sheet[f'H{current_row + 2}'] = canopy.get('section', '')
                    current_sheet[f'I{current_row + 2}'] = canopy.get('flowrate', '')
                    
                    # Set lights and quantities - default to "LIGHT SELECTION"
                    lights = canopy.get('lights', '')
                    if 'LED Strip' in lights:
                        current_sheet[f'C{current_row + 3}'] = "LED STRIP"
                    else:
                        current_sheet[f'C{current_row + 3}'] = "LIGHT SELECTION"  # Default value at C15
                    current_sheet[f'D{current_row + 3}'] = canopy.get('light_quantity', '')
                    
                    # Add light indicator in column C
                    if canopy.get('lights') or canopy.get('light_type'):
                        light_indicator_row = current_row + 11
                        current_sheet[f'C{light_indicator_row}'] = 1
                    
                    # Set special works
                    special_works_dict = canopy.get('specialWorks', {})
                    special_works_items = list(special_works_dict.items())[:2]
                    
                    # First special work
                    if len(special_works_items) > 0:
                        work, qty = special_works_items[0]
                        current_sheet[f'C{current_row + 4}'] = work
                        current_sheet[f'D{current_row + 4}'] = qty
                    
                    # Second special work
                    if len(special_works_items) > 1:
                        work, qty = special_works_items[1]
                        current_sheet[f'C{current_row + 5}'] = work
                        current_sheet[f'D{current_row + 5}'] = qty if qty else 1
                    else:
                        current_sheet[f'C{current_row + 5}'] = 'SELECT WORKS'
                    
                    # Wall cladding
                    if canopy.get('wallCladding'):
                        current_sheet[f'C{current_row + 7}'] = '2M² (HFL)'  # Always write 2M² (HFL) when cladding is selected
                        #current_sheet[f'D{current_row + 7}'] = canopy.get('cladding_length', 0)  # Length at D19
                        #current_sheet[f'E{current_row + 7}'] = canopy.get('cladding_height', 0)  # Height at E19
                        
                        # Create description string from selected walls
                        wall_desc = []
                        if 'Rear' in canopy.get('cladding_desc', []):
                            wall_desc.append('Rear')
                        if 'Left' in canopy.get('cladding_desc', []):
                            wall_desc.append('Left')
                        if 'Right' in canopy.get('cladding_desc', []):
                            wall_desc.append('Right')
                        
                        if wall_desc:
                            if len(wall_desc) > 1:
                                desc = f"Cladding to {', '.join(wall_desc[:-1])} & {wall_desc[-1]}-hand Walls"
                            else:
                                desc = f"Cladding to {wall_desc[0]}-hand Wall"
                            current_sheet[f'F{current_row + 7}'] = desc  # Description at F19
                    
                    # CMWI/CMWF specific fields
                    if canopy.get('model') in ['CMWI', 'CMWF']:
                        current_sheet[f'C{current_row + 13}'] = canopy.get('control_panel', '')
                        current_sheet[f'C{current_row + 14}'] = canopy.get('WW_pods', '')
                        current_sheet[f'D{current_row + 14}'] = canopy.get('WW_pods_quantity', '')
                        current_sheet[f'C{current_row + 15}'] = canopy.get('pipework', '')
                    
                    # If model is CXW, set extract static to 50 at F22 (first canopy)
                    if canopy.get('model') == 'CXW':
                        current_sheet[f'F{current_row + 10}'] = 50  # F22 for first canopy (12 + 10)
                    
                    current_row += 17
                
                # Add dropdowns to the sheet
                add_dropdowns_to_sheet(wb, current_sheet, 12)
                
                # After filling all data in the sheet, get the totals
                cost_cell = current_sheet['K9']  # Total cost from K9
                selling_price_cell = current_sheet['N9']  # Total selling price from N9
                
                # Add to running totals (handle None or error values)
                try:
                    if isinstance(cost_cell.value, (int, float)):
                        total_cost += float(cost_cell.value)
                except (TypeError, ValueError):
                    pass
                
                try:
                    if isinstance(selling_price_cell.value, (int, float)):
                        total_selling_price += float(selling_price_cell.value)
                except (TypeError, ValueError):
                    pass
                
                sheet_count += 1
        
        # Write totals to JOB TOTAL sheet
        if 'JOB TOTAL' in wb.sheetnames:
            job_total_sheet = wb['JOB TOTAL']
            job_total_sheet['S16'] = total_cost  # Write total cost to S16
            job_total_sheet['T16'] = total_selling_price  # Write total selling price to T16
        
        # Remove any unused CANOPY sheets
        for unused_sheet in canopy_sheets[sheet_count:]:
            del wb[unused_sheet]
        
        # Save workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        return output

    except Exception as e:
        st.error(f"Error generating Excel sheet: {str(e)}")
        return None

def generate_excel(kitchen_info, genInfo):
    """
    Generate an Excel file with canopy data.
    
    Args:
        kitchen_info (dict): Dictionary containing kitchen and canopy data.
        genInfo (dict): General information for the document.
        
    Returns:
        BytesIO: Excel file as a BytesIO object.
    """
    # Load the template
    template_path = TEMPLATES['EXCEL']
    wb = load_workbook(template_path)
    
    # Create a new workbook for each kitchen and floor
    for kitchen in kitchen_info:
        kitchen_name = kitchen['kitchen_name']
        
        for floor in kitchen['floors']:
            floor_name = floor['floor_name']
            sheet_name = f"CANOPY - {kitchen_name[:8]} {floor_name[:8]}"
            
            # Copy the template sheet
            template_sheet = wb['CANOPY']
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name[:31]  # Excel limits sheet names to 31 chars
            
            # Set header information
            new_sheet['C3'] = genInfo.get('projectName', '')
            new_sheet['C4'] = genInfo.get('projectNum', '')
            new_sheet['C5'] = genInfo.get('projectManager', '')
            new_sheet['C6'] = genInfo.get('salesPerson', '')
            new_sheet['C7'] = f"{kitchen_name} - {floor_name}"
            
            # Set canopy data
            current_row = 22  # Starting row for canopy data
            
            # Define single color for E and F columns
            text_color = '0000FF'  # Blue - or whichever single color you want
            
            for canopy in floor['canopies']:
                # Set cell values
                new_sheet[f'C{current_row}'] = canopy.get('itemNum', '')
                new_sheet[f'D{current_row}'] = canopy.get('model', '')
                new_sheet[f'E{current_row}'] = canopy.get('width', 0)
                new_sheet[f'F{current_row}'] = canopy.get('length', 0)

                # Set font color for E and F for both current row and row below
                for col in ['E', 'F']:
                    # Set font color for current row and row below
                    for row_offset in [0, 1]:
                        cell = new_sheet[f'{col}{current_row + row_offset}']
                        # Preserve existing font properties, only change color
                        old_font = cell.font
                        new_font = Font(
                            name=old_font.name,
                            size=old_font.size,
                            bold=old_font.bold,
                            italic=old_font.italic,
                            vertAlign=old_font.vertAlign,
                            underline=old_font.underline,
                            strike=old_font.strike,
                            color=text_color  # Use the single consistent color
                        )
                        cell.font = new_font

                # Set other canopy data
                new_sheet[f'G{current_row}'] = canopy.get('height', 0)
                new_sheet[f'H{current_row}'] = canopy.get('section', 0)
                new_sheet[f'I{current_row}'] = canopy.get('flowrate', 0)
                
                # Move to next canopy row
                current_row += 17  # Each canopy takes 17 rows
    
    # Remove the template sheet
    wb.remove(wb['CANOPY'])
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def apply_cell_formatting(cell):
    """Apply standard cell formatting for currency values"""
    # Light blue background - exact color from image
    cell.fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    
    # Thin borders on all sides
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border
    
    # Calibri font
    cell.font = Font(name='Calibri', size=11)
    
    # Currency format with 2 decimal places
    cell.number_format = '£#,##0.00'
    
    # Alignment
    cell.alignment = Alignment(horizontal='right')

def format_currency_cells(worksheet, current_row):
    """Apply currency formatting to specific cells"""
    # List of cells that need currency formatting
    currency_cells = [
        'K9', 'N9',  # Top cells
        f'K{current_row}', f'N{current_row}',  # Current row cells
        'P182'  # Bottom cell
    ]
    
    for cell_ref in currency_cells:
        cell = worksheet[cell_ref]
        apply_cell_formatting(cell)

def apply_value_cell_formatting(cell, add_border=False):
    """Apply formatting for value cells in J, K, N, and O columns"""
    # Only add border if specified (for J column)
    if add_border:
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # Bold Calibri font in dark blue
    cell.font = Font(name='Calibri', size=11, bold=True, color='000080')  # 000080 is dark blue
    
    # Currency format with 2 decimal places
    cell.number_format = '£#,##0.00'
    
    # Right alignment
    cell.alignment = Alignment(horizontal='right')

def format_value_cells(worksheet, start_row=14):
    """Format the J, K, N, and O columns for the canopy section"""
    # Format J14-J27, K14-K27, N14-N27, O14-O27
    for row in range(start_row, start_row + 14):  # 14 rows from start_row
        # Format J column with border
        j_cell = worksheet[f'J{row}']
        # Always apply font formatting
        j_cell.font = Font(name='Calibri', size=11, bold=True, color='9FCBDA')  # Changed color to #9FCBDA
        j_cell.number_format = '£#,##0.00'
        j_cell.alignment = Alignment(horizontal='right')
        # Add border for J column
        j_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format K, N, O columns without border
        for col in ['K', 'N', 'O']:
            cell = worksheet[f'{col}{row}']
            # Always apply font formatting
            cell.font = Font(name='Calibri', size=11, bold=True, color='9FCBDA')  # Changed color to #9FCBDA
            cell.number_format = '£#,##0.00'
            cell.alignment = Alignment(horizontal='right')

def get_wall_cladding(kitchen_info):
    """
    Extracts wall cladding information only for canopies that have cladding.
    Returns None if no canopies have wall cladding.
    """
    kitchens = kitchen_info.get("kitchens", [])
    wall_cladding_data = []

    for kitchen in kitchens:
        for floor in kitchen["floors"]:
            for canopy in floor["canopies"]:
                # Only process canopies that have wallCladding set to True and have cladding_desc
                if canopy.get('wallCladding') and canopy.get('cladding_desc'):
                    # Get the canopy's item number
                    item_num = canopy.get('itemNum') or canopy.get('item_number', '1')
                    
                    # Get the selected walls
                    selected_walls = canopy.get('cladding_desc', [])
                    
                    if selected_walls:
                        wall_parts = []
                        if 'Rear' in selected_walls:  # Put Rear first
                            wall_parts.append('Rear')
                        if 'Left' in selected_walls:
                            wall_parts.append('Left')
                        if 'Right' in selected_walls:
                            wall_parts.append('Right')
                        
                        if wall_parts:
                            # Format as "Rear, Left & Right-hand Walls"
                            if len(wall_parts) > 1:
                                wall_description = f"Cladding to {', '.join(wall_parts[:-1])} & {wall_parts[-1]}-hand Walls"
                            else:
                                wall_description = f"Cladding to {wall_parts[0]}-hand Wall"
                        else:
                            continue  # Skip if no walls selected
                    else:
                        continue  # Skip if no cladding description

                    # Get cladding price and round up to nearest integer
                    cladding_price = canopy.get('cladding_price', 0)
                    if isinstance(cladding_price, (int, float)):
                        cladding_price = math.ceil(cladding_price)

                    wall_cladding_data.append({
                        "item_no": str(item_num),  # Use the canopy's item number
                        "description": wall_description,
                        "width": canopy.get("cladding_width", 0),
                        "height": canopy.get("cladding_height", 0),
                        "price": cladding_price
                    })

    # Only return data if we found canopies with cladding
    return wall_cladding_data if wall_cladding_data else None

def calculate_cladding_subtotal(canopies):
    """Calculate total cladding price for a floor's canopies"""
    cladding_total = 0
    for canopy in canopies:
        if canopy.get('wallCladding'):
            cladding_price = float(canopy.get('cladding_price', 0))
            cladding_total += cladding_price
    return round(cladding_total, 2)

def calculate_floor_subtotal(floor_canopies, delivery_install_data, commission_price=0):
    """Calculate subtotal for a floor including all canopies, P182, cladding, delivery/installation, and commissioning"""
    subtotal = 0
    
    # Sum all canopy prices and P182
    for canopy in floor_canopies:
        canopy_price = float(canopy.get('total_price', 0))
        p182_value = float(canopy.get('p182_value', 0))
        subtotal += canopy_price + p182_value
    
    # Calculate cladding total
    cladding_total = calculate_cladding_subtotal(floor_canopies)
    subtotal += cladding_total
    
    # Add delivery and installation costs
    delivery_price = float(delivery_install_data.get('delivery_price', 0))
    install_price = float(delivery_install_data.get('install_price', 0))
    
    # Add commissioning price
    subtotal += delivery_price + install_price + commission_price
    
    return round(subtotal, 2), round(cladding_total, 2)

def add_dropdowns_to_sheet(wb, sheet, start_row):
    """Add data validation (dropdowns) to specific cells"""
    if 'Lists' not in wb.sheetnames:
        list_sheet = wb.create_sheet('Lists')
    else:
        list_sheet = wb['Lists']

    # Define all dropdown options
    dropdowns = {
        'lights': {
            'options': [
                'LED STRIP L6 Inc DALI',
                'LED STRIP L12 inc DALI', 
                'LED STRIP L18 Inc DALI',
                'Small LED Spots inc DALI',
                'LARGE LED Spots inc DALI'
            ],
            'column': 'A',
            'target_col': 'C',
            'row_offset': 3  # C15 for first canopy
        },
        'special_works_1': {
            'options': [
                'ROUND CORNERS',
                'CUT OUT',
                'CASTELLE LOCKING',
                'HEADER DUCT S/S',
                'HEADER DUCT',
                'PAINT FINSH',
                'UV ON DEMAND',
                'E/over for emergency strip light',
                'E/over for small emer. spot light',
                'E/over for large emer. spot light',
                'COLD MIST ON DEMAND',
                'CMW  PIPEWORK HWS/CWS',
                'CANOPY GROUND SUPPORT',
                '2nd EXTRACT PLENUM',
                'SUPPLY AIR PLENUM',
                'CAPTUREJET PLENUM',
                'COALESCER'
            ],
            'column': 'B',
            'target_col': 'C',
            'row_offset': 4  # C16 for first special work
        },
        'special_works_2': {
            'options': [
                'SELECT WORKS',  # Add default option
                'ROUND CORNERS',
                'CUT OUT',
                'CASTELLE LOCKING',
                'HEADER DUCT S/S',
                'HEADER DUCT',
                'PAINT FINSH',
                'UV ON DEMAND',
                'E/over for emergency strip light',
                'E/over for small emer. spot light',
                'E/over for large emer. spot light',
                'COLD MIST ON DEMAND',
                'CMW  PIPEWORK HWS/CWS',
                'CANOPY GROUND SUPPORT',
                '2nd EXTRACT PLENUM',
                'SUPPLY AIR PLENUM',
                'CAPTUREJET PLENUM',
                'COALESCER'
            ],
            'column': 'B',  # Use same column as they're the same options
            'target_col': 'C',
            'row_offset': 5  # C17 for second special work
        },
        'wall_cladding': {
            'options': ['', '2M² (HFL)'],
            'column': 'C',
            'target_col': 'C',
            'row_offset': 7  # C19 for first canopy (12 + 7 = 19)
        },
        'control_panel': {
            'options': ['CP1S', 'CP2S', 'CP3S', 'CP4S'],
            'column': 'D',
            'target_col': 'C',
            'row_offset': 13  # C25 for first canopy
        },
        'ww_pods': {
            'options': [
                '1000-S', '1500-S', '2000-S', '2500-S', '3000-S',
                '1000-D', '1500-D', '2000-D', '2500-D', '3000-D'
            ],
            'column': 'E',
            'target_col': 'C',
            'row_offset': 14  # C26 for first canopy
        },
        'delivery_location': {
            'options': [
                "",  # Empty option first
                "ABERDEEN 590",
                "ABINGDON 110",
                "ALDEBURGH 112",
                "ALDERSHOT 110",
                "ALNWICK 342",
                "ANDOVER 110",
                "ASHFORD 25",
                "AYLESBURY 86",
                "BANBURY 102",
                "BANGOR 324",
                "BARKING 32",
                "BARNET 55",
                "BARNSLEY 209",
                "BARNSTABLE 227",
                "BARROW-IN-FURNESS 348",
                "BASILDON 38",
                "BASINGSTOKE 82",
                "BATH 154",
                "BEDFORD 103",
                "BERWICK-UPON-TWEED 371",
                "BILLERICAY 37",
                "BIRKENHEAD 277",
                "BIRMINGHAM 168",
                "BLACKBURN 283",
                "BLACKPOOL 289",
                "BLANDFORD FORUM 144",
                "BODMIN 273",
                "BOGNOR REGIS 88",
                "BOLTON 259",
                "BOOTLE 272",
                "BOURNEMOUTH 140",
                "BRADFORD 234",
                "BRAINTREE 60",
                "BRIDGEND 205",
                "BRIDLINGTON 244",
                "BRIGHTON 68",
                "BRISTOL 157",
                "BUCKINGHAMSHIRE 109",
                "BURNLEY 296",
                "BURTON UPON TRENT 175",
                "BURY ST EDMUNDS 98",
                "CAMBRIDGE 85",
                "CANNOCK 175",
                "CANTERBURY 30",
                "CARDIFF 192",
                "CARLISLE 356",
                "CARMARTHEN 252",
                "CHELTENHAM 148",
                "CHESTER 268",
                "COVENTRY 146",
                "CHIPPENHAM 136",
                "COLCHESTER 78",
                "CORBY 128",
                "DARTMOUTH 245",
                "DERBY 178",
                "DONCASTER 203",
                "DORCHESTER 160",
                "DORKING 46",
                "DOVER 45",
                "DURHAM 299",
                "EASTBOURNE 57",
                "EASTLEIGH 109",
                "EDINBURGH 428",
                "ENFIELD 49",
                "EXETER 205",
                "EXMOUTH 207",
                "FELIXSTOWE 103",
                "GATWICK 44",
                "GLASGOW 456",
                "GLASTONBURY 164",
                "GLOUCESTER 151",
                "GRANTHAM 143",
                "GREAT YARMOUTH 147",
                "GRIMSBY 215",
                "GUILDFORD 59",
                "HARLOW 47",
                "HARROGATE 236",
                "HARTLEPOOL 286",
                "HASTINGS 40",
                "HEXHAM 325",
                "HEREFORD 184",
                "HIGH WYCOMBE 80",
                "HIGHBRIDGE 187",
                "HONITON 190",
                "HORSHAM 55",
                "HOUNSLOW 55",
                "HUDDERSFIELD 239",
                "HULL 247",
                "HUNTINGDON 94",
                "INVERNESS 619",
                "IPSWICH 94",
                "IRELAND",
                "KENDAL 321",
                "KETTERING 127",
                "KIDDERMINSTER 179",
                "KILMARNOCK 449",
                "KINGSTON UPON HULL 220",
                "KINGSTON UPON THAMES 52",
                "LANCASTER 290",
                "LAUNCESTON 251",
                "LEAMINGTON SPA 146",
                "LEEDS 231",
                "LEICESTER 151",
                "LEIGH ON SEA 45",
                "LEWISHAM 29",
                "LINCOLN 179",
                "LIVERPOOL 258",
                "LLANDUDNO 309",
                "LONDON in FORS GOLD(varies)",
                "LUTON 80",
                "MABLETHORPE 182",
                "MACCLESFIELD 244",
                "MANCHESTER 251",
                "MARGATE 46",
                "MIDDLESBROUGH 286",
                "MILFORD HAVEN 289",
                "MILTON KEYNES 101",
                "MORPETH 327",
                "NANTWICH 232",
                "NEWBURY 101",
                "NEWCASTLE 308",
                "NEWPORT 178",
                "NEWQUAY 178",
                "NORTHAMPTON 116",
                "NORTHUMBERLAND 341",
                "NORWICH 136",
                "NOTTINGHAM 177",
                "OKEHAMPTON 232",
                "OXFORD 106",
                "PENRITH 316",
                "PENZANCE 318",
                "PERTH 477",
                "PETERBOROUGH 124",
                "PETERSFIELD 87",
                "PETWORTH 71",
                "PLYMOUTH 247",
                "PONTEFRACT 221",
                "POOLE 144",
                "PORTSMOUTH 102",
                "READING 88",
                "REIGATE 39",
                "RINGWOOD 130",
                "ROSS-ON-WYE 171",
                "ROTHERHAM 203",
                "SALISBURY 120",
                "SCARBOROUGH 277",
                "SCUNTHORPE 204",
                "SHEFFIELD 205",
                "SHREWSBURY 207",
                "SHROPSHIRE 218",
                "SLOUGH 72",
                "SOUTH SHIELDS 310",
                "SOUTHAMPTON 112",
                "SOUTHEND 52",
                "SOUTHPORT 279",
                "SPALDING 143",
                "ST ALBANS 62",
                "ST IVES 317",
                "STAFFORD 187",
                "STAINES 61",
                "STEVENAGE 72",
                "STIRLING 445",
                "STOCKPORT 257",
                "STOCKTON 278",
                "STOKE-ON-TRENT 205",
                "STRATFORD UPON AVON 151",
                "SUNDERLAND 309",
                "SWINDON 121",
                "TAMWORTH 180",
                "TAUNTON 185",
                "TELFORD 193",
                "TILBURY 34",
                "TORQUAY 227",
                "TUNBRIDGE WELLS 26",
                "UXBRIDGE 74",
                "WAKEFIELD 214",
                "WARMISTER 137",
                "WARWICK 148",
                "WATFORD 67",
                "WELSHPOOL 238",
                "WEMBLEY 55",
                "WEYMOUTH 173",
                "WHITBY 282",
                "WIGAN 252",
                "WINCANTON 149",
                "WINCHESTER 100",
                "WOKING 60",
                "WOLVERHAMPTON 175",
                "WORCESTER 160",
                "WREXHAM 250",
                "YEOVIL 163",
                "YORK 243"
            ],
            'column': 'F',
            'target_col': 'D',
            'target_row': 183  # Fixed row for delivery location
        },
        'plant_hire_1': {
            'options': [
                "",
                "SL10 GENIE",
                "EXTENSION FORKS",
                "2.5M COMBI LADDER",
                "1.5M PODIUM",
                "3M TOWER",
                "COMBI LADDER",
                "PECO LIFT",
                "3M YOUNGMAN BOARD",
                "GS1930 SCISSOR LIFT",
                "4-6 SHERASCOPIC",
                "7-9 SHERASCOPIC"
            ],
            'column': 'G',
            'target_col': 'D',
            'target_row': 184  # Fixed row for first plant hire
        },
        'plant_hire_2': {
            'options': [
                "",
                "SL10 GENIE",
                "EXTENSION FORKS",
                "2.5M COMBI LADDER",
                "1.5M PODIUM",
                "3M TOWER",
                "COMBI LADDER",
                "PECO LIFT",
                "3M YOUNGMAN BOARD",
                "GS1930 SCISSOR LIFT",
                "4-6 SHERASCOPIC",
                "7-9 SHERASCOPIC"
            ],
            'column': 'G',  # Can use same column as plant_hire_1
            'target_col': 'D',
            'target_row': 185  # Fixed row for second plant hire
        }
    }

    # Write options to Lists sheet and create validations
    for name, config in dropdowns.items():
        # Write options to Lists sheet
        for i, option in enumerate(config['options'], 1):
            list_sheet[f"{config['column']}{i}"] = option
        
        # Create range reference
        range_ref = f"Lists!${config['column']}$1:${config['column']}${len(config['options'])}"
        
        # Create validation
        dv = DataValidation(
            type="list",
            formula1=range_ref,
            allow_blank=True
        )
        sheet.add_data_validation(dv)
        
        if 'target_row' in config:
            # Fixed position dropdown (delivery/installation)
            dv.add(f"{config['target_col']}{config['target_row']}")
        else:
            # Repeating dropdowns (canopy-related)
            current_row = start_row + config['row_offset']
            while current_row <= sheet.max_row:
                dv.add(f"{config['target_col']}{current_row}")
                current_row += 17

def generate_initial_excel(kitchen_info, genInfo):
    """Generate Excel with basic sheets and minimal info"""
    try:
        excel_path = TEMPLATES['EXCEL']
        wb = load_workbook(excel_path, data_only=False)  # Keep formulas
        template_ws = wb['CANOPY']
        
        # Create sheets for each floor
        for kitchen in kitchen_info:
            kitchen_name = kitchen['kitchen_name']
            for floor in kitchen['floors']:
                floor_name = floor['floor_name']
                sheet_name = f"CANOPY - {floor_name} ({kitchen_name})"[:31]
                
                # Create new sheet from template
                new_sheet = wb.copy_worksheet(template_ws)
                new_sheet.title = sheet_name
                
                # Write general info
                new_sheet['C3'] = genInfo.get('projectNum', '')
                new_sheet['C5'] = genInfo.get('customer', '')
                new_sheet['C7'] = genInfo.get('combined_initials', '')
                new_sheet['G3'] = genInfo.get('projectName', '')
                new_sheet['G5'] = genInfo.get('location', '')
                new_sheet['G7'] = genInfo.get('date', '')
                
                # Write canopy info
                current_row = 12
                for canopy in floor.get('canopies', []):
                    try:
                        # Write basic canopy info
                        new_sheet[f'B{current_row}'] = canopy.get('itemNum', '')
                        new_sheet[f'C{current_row + 2}'] = canopy.get('configuration', '')
                        new_sheet[f'D{current_row + 2}'] = canopy.get('model', '')
                            
                        # Handle cladding dimensions if wallCladding is True
                        if canopy.get('wallCladding'):
                            length = canopy.get('length', 0)
                            height = canopy.get('height', 0)
                            if length and height:
                                # Write dimensions in format "LengthxHeight" to G19
                                cladding_dimensions = f"{length}x{height}"
                                new_sheet[f'G{current_row + 7}'] = cladding_dimensions
                                    
                    except Exception as e:
                        st.error(f"Error writing canopy data at row {current_row}: {str(e)}")
                    
                    current_row += 17  # Move to next canopy section
        
        # Remove template sheet if it exists
        if 'CANOPY' in wb.sheetnames:
            wb.remove(wb['CANOPY'])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    except Exception as e:
        st.error(f"Error generating Excel sheet: {str(e)}")
        return None