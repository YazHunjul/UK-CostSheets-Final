import streamlit as st
import zipfile
from io import BytesIO
from costSheetGen.Canopy import canopyExcel as CE
from costSheetGen.Canopy import canopyWord as CW
import time
from openpyxl import load_workbook
import io
import os
from costSheetGen.Canopy.canopyUtils import extract_canopy_prices, convert_formulas_to_values, run_excel_script
import pyperclip
import pandas as pd
import openpyxl
import math
from ..config import TEMPLATES
import json
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation

canopyWithCladding =[]
def get_initials(name):
    """Extract initials from a name"""
    if name:
        words = name.split()
        initials = ''.join(word[0].upper() for word in words if word)
        return initials
    return ''

def get_delivery_install_details(floor_name, key_suffix):
    """
    Collect delivery and installation details for a specific floor
    
    Args:
        floor_name: Name of the floor for labeling
        key_suffix: Unique suffix for Streamlit keys to prevent conflicts
    """
    with st.expander(f"🚚 Delivery & Installation Details - {floor_name}", expanded=False):
        st.markdown('<div class="section-header"><h3>📍 Location & Plant Hire</h3></div>', unsafe_allow_html=True)
        
        col1, col2 = st.columns([1, 3])
        with col1:
            delivery_lift_qty = st.number_input(
                "DELIVERY 1 x 7.5T TAIL LIFT",
                min_value=0,
                value=1,
                step=1,
                key=f"delivery_lift_qty_{key_suffix}"
            )

        with col2:
            delivery_locations = [
                "",
                "ABERDEEN 590",
                "ABINGDON 110",
                "ALDEBURGH 112",
                "ALDERSHOT 110",
                "ALNWICK 342",
                "ANDOVER 110",
                "ASHFORD 25",
                "AYLESBURY 86",
                "BANBURY 102",
                "BANGOR 324",
                "BARKING 32",
                "BARNET 55",
                "BARNSLEY 209",
                "BARNSTABLE 227",
                "BARROW-IN-FURNESS 348",
                "BASILDON 38",
                "BASINGSTOKE 82",
                "BATH 154",
                "BEDFORD 103",
                "BERWICK-UPON-TWEED 371",
                "BILLERICAY 37",
                "BIRKENHEAD 277",
                "BIRMINGHAM 168",
                "BLACKBURN 283",
                "BLACKPOOL 289",
                "BLANDFORD FORUM 144",
                "BODMIN 273",
                "BOGNOR REGIS 88",
                "BOLTON 259",
                "BOOTLE 272",
                "BOURNEMOUTH 140",
                "BRADFORD 234",
                "BRAINTREE 60",
                "BRIDGEND 205",
                "BRIDLINGTON 244",
                "BRIGHTON 68",
                "BRISTOL 157",
                "BUCKINGHAMSHIRE 109",
                "BURNLEY 296",
                "BURTON UPON TRENT 175",
                "BURY ST EDMUNDS 98",
                "CAMBRIDGE 85",
                "CANNOCK 175",
                "CANTERBURY 30",
                "CARDIFF 192",
                "CARLISLE 356",
                "CARMARTHEN 252",
                "CHELTENHAM 148",
                "CHESTER 268",
                "COVENTRY 146",
                "CHIPPENHAM 136",
                "COLCHESTER 78",
                "CORBY 128",
                "DARTMOUTH 245",
                "DERBY 178",
                "DONCASTER 203",
                "DORCHESTER 160",
                "DORKING 46",
                "DOVER 45",
                "DURHAM 299",
                "EASTBOURNE 57",
                "EASTLEIGH 109",
                "EDINBURGH 428",
                "ENFIELD 49",
                "EXETER 205",
                "EXMOUTH 207",
                "FELIXSTOWE 103",
                "GATWICK 44",
                "GLASGOW 456",
                "GLASTONBURY 164",
                "GLOUCESTER 151",
                "GRANTHAM 143",
                "GREAT YARMOUTH 147",
                "GRIMSBY 215",
                "GUILDFORD 59",
                "HARLOW 47",
                "HARROGATE 236",
                "HARTLEPOOL 286",
                "HASTINGS 40",
                "HEXHAM 325",
                "HEREFORD 184",
                "HIGH WYCOMBE 80",
                "HIGHBRIDGE 187",
                "HONITON 190",
                "HORSHAM 55",
                "HOUNSLOW 55",
                "HUDDERSFIELD 239",
                "HULL 247",
                "HUNTINGDON 94",
                "INVERNESS 619",
                "IPSWICH 94",
                "IRELAND",
                "KENDAL 321",
                "KETTERING 127",
                "KIDDERMINSTER 179",
                "KILMARNOCK 449",
                "KINGSTON UPON HULL 220",
                "KINGSTON UPON THAMES 52",
                "LANCASTER 290",
                "LAUNCESTON 251",
                "LEAMINGTON SPA 146",
                "LEEDS 231",
                "LEICESTER 151",
                "LEIGH ON SEA 45",
                "LEWISHAM 29",
                "LINCOLN 179",
                "LIVERPOOL 258",
                "LLANDUDNO 309",
                "LONDON in FORS GOLD (varies)",
                "LUTON 80",
                "MABLETHORPE 182",
                "MACCLESFIELD 244",
                "MANCHESTER 251",
                "MARGATE 46",
                "MIDDLESBROUGH 286",
                "MILFORD HAVEN 289",
                "MILTON KEYNES 101",
                "MORPETH 327",
                "NANTWICH 232",
                "NEWBURY 101",
                "NEWCASTLE 308",
                "NEWPORT 178",
                "NEWQUAY 178",
                "NORTHAMPTON 116",
                "NORTHUMBERLAND 341",
                "NORWICH 136",
                "NOTTINGHAM 177",
                "OKEHAMPTON 232",
                "OXFORD 106",
                "PENRITH 316",
                "PENZANCE 318",
                "PERTH 477",
                "PETERBOROUGH 124",
                "PETERSFIELD 87",
                "PETWORTH 71",
                "PLYMOUTH 247",
                "PONTEFRACT 221",
                "POOLE 144",
                "PORTSMOUTH 102",
                "READING 88",
                "REIGATE 39",
                "RINGWOOD 130",
                "ROSS-ON-WYE 171",
                "ROTHERHAM 203",
                "SALISBURY 120",
                "SCARBOROUGH 277",
                "SCUNTHORPE 204",
                "SHEFFIELD 205",
                "SHREWSBURY 207",
                "SHROPSHIRE 218",
                "SLOUGH 72",
                "SOUTH SHIELDS 310",
                "SOUTHAMPTON 112",
                "SOUTHEND 52",
                "SOUTHPORT 279",
                "SPALDING 143",
                "ST ALBANS 62",
                "ST IVES 317",
                "STAFFORD 187",
                "STAINES 61",
                "STEVENAGE 72",
                "STIRLING 445",
                "STOCKPORT 257",
                "STOCKTON 278",
                "STOKE-ON-TRENT 205",
                "STRATFORD UPON AVON 151",
                "SUNDERLAND 309",
                "SWINDON 121",
                "TAMWORTH 180",
                "TAUNTON 185",
                "TELFORD 193",
                "TILBURY 34",
                "TORQUAY 227",
                "TUNBRIDGE WELLS 26",
                "UXBRIDGE 74",
                "WAKEFIELD 214",
                "WARMISTER 137",
                "WARWICK 148",
                "WATFORD 67",
                "WELSHPOOL 238",
                "WEMBLEY 55",
                "WEYMOUTH 173",
                "WHITBY 282",
                "WIGAN 252",
                "WINCANTON 149",
                "WINCHESTER 100",
                "WOKING 60",
                "WOLVERHAMPTON 175",
                "WORCESTER 160",
                "WREXHAM 250",
                "YEOVIL 163",
                "YORK 243"
            ]
            
            delivery_location = st.selectbox(
                "SELECT LOCATION",
                options=delivery_locations,
                help="Select delivery location",
                key=f"delivery_location_{key_suffix}"
            )

        # Plant Hire Section
        plant_hires = st.multiselect(
            "Select Plant Hires (max 2)",
            options=["Plant Hire 1", "Plant Hire 2"],
            max_selections=2,
            key=f"plant_hires_{key_suffix}"
        )

        plant_selections = {}
        quantities = {}
        for plant in plant_hires:
            col1, col2 = st.columns([3, 1])
            with col1:
                plant_selections[plant] = st.selectbox(
                    f"PLANT SELECTION (weekly) for {plant}",
                    ["", "SL10 GENIE", "EXTENSION FORKS", "2.5M COMBI LADDER", 
                     "1.5M PODIUM", "3M TOWER", "COMBI LADDER", "PECO LIFT", 
                     "3M YOUNGMAN BOARD", "GS1930 SCISSOR LIFT", "4-6 SHERASCOPIC", 
                     "7-9 SHERASCOPIC"],
                    key=f"plant_selection_{plant}_{key_suffix}"
                )
            
            with col2:
                if plant_selections[plant]:
                    quantities[plant] = st.number_input(
                        "QTY",
                        min_value=1,
                        value=1,
                        step=1,
                        key=f"qty_{plant}_{key_suffix}"
                    )

        # Installation details in columns
        col1, col2 = st.columns(2)
        
        with col1:
            strip_out = st.number_input("STRIP OUT (PER DAY)", min_value=0.0, value=0.0, step=1.0, key=f"strip_out_{key_suffix}")
            #consumables = st.number_input("CONSUMABLES 15(P) + 19(H)", min_value=0.0, value=0.0, step=1.0, key=f"consumables_{key_suffix}")
            installation_normal = st.number_input("INSTALLATION NORMAL HOURS", min_value=0.0, value=0.0, step=1.0, key=f"installation_normal_{key_suffix}")
            installation_after = st.number_input("INSTALLATION AFTER HOURS", min_value=0.0, value=0.0, step=1.0, key=f"installation_after_{key_suffix}")
            wall_cladding = st.number_input("WALL CLADDING INSTALLATION", min_value=0.0, value=0.0, step=1.0, key=f"wall_cladding_{key_suffix}")
            
        with col2:
            overnight_expenses = st.number_input("OVERNIGHT/TRAVEL EXPENSES", min_value=0.0, value=0.0, step=1.0, key=f"overnight_expenses_{key_suffix}")
            test_commission = st.number_input("TEST & COMMISSION", min_value=0.0, value=0.0, step=1.0, key=f"test_commission_{key_suffix}")
            gas_interlock = st.number_input("GAS INTERLOCK (INSTALLED)", min_value=0.0, value=0.0, step=1.0, key=f"gas_interlock_{key_suffix}")
            co_sensor = st.number_input("CO SENSOR (SOLID FUEL)", min_value=0.0, value=0.0, step=1.0, key=f"co_sensor_{key_suffix}")
            co2_sensor = st.number_input("CO2 SENSOR (DCK)", min_value=0.0, value=0.0, step=1.0, key=f"co2_sensor_{key_suffix}")
            bms_interface = st.number_input("BMS FAULT INTERFACE", min_value=0.0, value=0.0, step=1.0, key=f"bms_interface_{key_suffix}")

        return {
            "delivery_location": delivery_location,
            "delivery_lift_qty": delivery_lift_qty,
            "plant_hires": plant_selections,
            "quantities": quantities,
            "strip_out": strip_out,
            # "consumables": consumables,
            "installation_normal": installation_normal,
            "installation_after": installation_after,
            "wall_cladding": wall_cladding,
            "overnight_expenses": overnight_expenses,
            "test_commission": test_commission,
            "gas_interlock": gas_interlock,
            "co_sensor": co_sensor,
            "co2_sensor": co2_sensor,
            "bms_interface": bms_interface,
        }

def load_session_state(json_data):
    """Load session state from JSON data"""
    try:
        data = json.loads(json_data)
        
        # Only restore session state values, skip general info
        if 'session_state' in data:
            # First clear existing session state (except system keys)
            keys_to_keep = {k for k in st.session_state.keys() if k.startswith('_')}
            for k in list(st.session_state.keys()):
                if k not in keys_to_keep and not k.startswith('customer'):  # Skip customer and other genInfo fields
                    del st.session_state[k]
            
            # Then restore session state values
            for key, value in data['session_state'].items():
                if not key.startswith('_') and not key.startswith('customer'):  # Skip system keys and customer fields
                    st.session_state[key] = value
        
        # Force a rerun to ensure all widgets update
        st.rerun()
        return True
        
    except Exception as e:
        st.error(f"Error loading project data: {str(e)}")
        return False

def save_session_state():
    """Save current session state to a JSON file"""
    # Get all relevant session state data
    save_data = {
        'session_state': {
            k: v for k, v in st.session_state.items() 
            if not k.startswith('_') and not k.startswith('customer')  # Exclude internal keys and customer fields
        },
        'timestamp': datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    }
    
    # Convert to JSON string
    json_str = json.dumps(save_data, indent=2)
    
    return json_str

def add_save_load_section():
    """Add save/load UI section"""
    with st.expander("💾 Save/Load Project Data"):
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### Save Project")
            if st.button("Generate Project File"):
                json_str = save_session_state()
                project_name = st.session_state.get('projectNum', 'project')
                
                st.download_button(
                    label="⬇️ Download Project File",
                    data=json_str,
                    file_name=f"{project_name}_data.json",
                    mime="application/json"
                )
        
        with col2:
            st.markdown("### Load Project")
            uploaded_file = st.file_uploader(
                "Upload Project File", 
                type=['json'],
                help="Upload a previously saved project file"
            )
            
            if uploaded_file is not None:
                json_str = uploaded_file.read().decode()
                if load_session_state(json_str):
                    st.success("Project data loaded successfully!")
                    st.rerun()  # Refresh the page to show loaded data

def main(genInfo):
    st.markdown('<hr>', unsafe_allow_html=True)
    st.markdown("<h2 style='text-align: center;'>Canopy Cost Sheet</h2>", unsafe_allow_html=True)
    
    # Add save/load section at the top
    add_save_load_section()
    
    st.markdown("---")
    
    # Add anchors for main sections
    st.markdown("<div id='general_info'></div>", unsafe_allow_html=True)
    
    # Get Kitchen Count first
    num_kitchens = st.number_input("Enter Number of Levels", min_value=1, key='num_kitchens')
    kitchen_info = []

    # Navigation sidebar
    with st.sidebar:
        st.markdown("### 🔍 Quick Navigation")
        
        # Main sections navigation
        st.markdown("#### Main Sections:")
        st.markdown("[📋 General Information](#general_info)")
        st.markdown("[💾 Generate Files](#generate)")
        
        # Floor-Area navigation
        st.markdown("#### Jump to Floor - Area:")
        
        for i in range(num_kitchens):
            kitchen_name = st.session_state.get(f'kitchen_name_{i}', f'Floor {i + 1}')
            num_floors = st.session_state.get(f'floors_input_{i}', 1)
            
            if i > 0:
                st.markdown("---")
            
            for floor in range(num_floors):
                floor_name = st.session_state.get(f'floor_name_{i}_{floor}', f'Area {floor + 1}')
                st.markdown(f"[{kitchen_name} - {floor_name}](#floor_{i}_{floor})")

    # Main form content
    for i in range(num_kitchens):
        st.markdown(f"<div id='kitchen_{i}'></div>", unsafe_allow_html=True)
        kitchen_name = st.text_input(f"Enter Level {i + 1} Name", key=f'kitchen_name_{i}')
        if kitchen_name:
            kitchen_data = {
                "kitchen_name": kitchen_name,
                "floors": []
            }

            with st.expander(f'{kitchen_name.title()} Floor Information', expanded=False):
                num_floors = st.number_input(
                    f"Enter the number of areas in {kitchen_name} Floor", 
                    min_value=1, 
                    key=f'floors_input_{i}'
                )
                for floor in range(num_floors):
                    # Add anchor for floor
                    st.markdown(f"<div id='floor_{i}_{floor}'></div>", unsafe_allow_html=True)
                    floor_name = st.text_input(
                        f"Enter area {floor + 1} Name", 
                        key=f'floor_name_{i}_{floor}'
                    )
                    if floor_name:
                        # Create a dictionary for this floor
                        floor_data = {
                            "floor_name": floor_name,
                            "canopies": []
                        }

                        num_canopies = st.number_input(
                            f"Enter Number of Canopies in {floor_name}",
                            min_value=1, 
                            key=f'canopies_input_{i}_{floor}'
                        )
                        for canopy in range(num_canopies):
                            st.markdown(f"<h4 style='text-align:center;'>Canopy {canopy + 1} - Floor: ({floor_name})</h4>", unsafe_allow_html=True)

                            coll1, coll2, coll3 = st.columns(3)
                            with coll1:
                                item_number = st.text_input('Reference Number', key=f'itemNum_{i}_{floor}_{canopy}')

                            with coll2:
                                configuration = st.selectbox('Configuration', ['WALL', "ISLAND"], key=f'config_{i}_{floor}_{canopy}')

                            with coll3:
                                model = st.selectbox(
                                    'Model', 
                                    ['KVF', 'KVX-M', "KVI", "UVX", "UVX-M", "UVI", "UVF", "UV-C POD", "CMWI", "CMWF", "CXW", "CXW-M", "KVV"], 
                                    key=f'model_{i}_{floor}_{canopy}'
                                )

                            # Initialize cladding variables
                            cladding_length = 0
                            cladding_height = 0
                            cladding_desc = []

                            # Add wall cladding selection
                            cladding_col1, cladding_col2, cladding_col3 = st.columns(3)
                            with cladding_col1:
                                wall_cladding = st.selectbox(
                                    'Wall Cladding',
                                    ['', '2M² (HFL)'],
                                    key=f'wall_cladding_{i}_{floor}_{canopy}'
                                )

                            st.write(f"Debug - Selected wall cladding: {wall_cladding}")  # Debug print

                            # Only show additional cladding fields if 2M² (HFL) is selected
                            if wall_cladding == '2M² (HFL)':
                                with cladding_col2:
                                    cladding_length = st.number_input(
                                        "Length",
                                        min_value=0,
                                        key=f'cladding_length_{i}_{floor}_{canopy}'
                                    )
                                with cladding_col3:
                                    cladding_height = st.number_input(
                                        "Height",
                                        min_value=0,
                                        key=f'cladding_height_{i}_{floor}_{canopy}'
                                    )
                                
                                # Wall selection
                                cladding_desc = st.multiselect(
                                    "Select Walls",
                                    options=["Rear", "Left", "Right"],
                                    key=f'cladding_desc_{i}_{floor}_{canopy}'
                                )

                            # Create canopy data dictionary
                            canopy_data = {
                                'item_number': item_number,
                                'model': model,
                                'configuration': configuration,
                                'section': 0,
                                'height': 0,
                                'width': 0,
                                'length': 0,
                                'lights': '',
                                'light_quantity': 0,
                                'flowrate': 0.0,
                                'specialWorks': {},
                                'control_panel': '',
                                'WW_pods': '',
                                'pipework': 0,
                                'WW_pods_quantity': 0,
                                'wallCladding': wall_cladding,  # Add wall cladding selection
                                'cladding_length': cladding_length,  # No need for conditional here
                                'cladding_height': cladding_height,
                                'cladding_desc': cladding_desc
                            }

                            st.write(f"Debug - Canopy data wall cladding: {canopy_data.get('wallCladding')}")  # Debug print

                            # Append canopy data to the floor
                            floor_data["canopies"].append(canopy_data)

                        # Append floor data to the kitchen
                        kitchen_data["floors"].append(floor_data)

            # Append kitchen data to the main list
            kitchen_info.append(kitchen_data)

    # Remove the Delivery & Installation Section
    # st.markdown("<div id='delivery'></div>", unsafe_allow_html=True)
    # st.markdown("## 🚚 Delivery & Installation Details")
    # ... remove the entire delivery and installation section ...

    st.markdown('<hr>', unsafe_allow_html=True)

    # Document generation section
    st.markdown("<div id='generate'></div>", unsafe_allow_html=True)
    st.markdown("## 💾 Generate Files")
    st.markdown("### Step 1: Generate and Download Excel")
    if st.button("Generate Excel File"):
        try:
            # Simplify the kitchen_info structure to only include essential data
            basic_kitchen_info = []
            for kitchen in kitchen_info:
                basic_kitchen = {
                    "kitchen_name": kitchen["kitchen_name"],
                    "floors": []
                }
                
                for floor in kitchen["floors"]:
                    basic_floor = {
                        "floor_name": floor["floor_name"],
                        "canopies": []
                    }
                    
                    for canopy in floor["canopies"]:
                        # Include all necessary canopy information
                        basic_canopy = {
                            "item_number": canopy.get("item_number", ""),
                            "model": canopy.get("model", ""),
                            "configuration": canopy.get("configuration", ""),
                            "wallCladding": canopy.get("wallCladding", ""),  # Add wall cladding
                            "cladding_length": canopy.get("cladding_length", 0),
                            "cladding_height": canopy.get("cladding_height", 0),
                            "cladding_desc": canopy.get("cladding_desc", [])
                        }
                        basic_floor["canopies"].append(basic_canopy)
                    
                    basic_kitchen["floors"].append(basic_floor)
                basic_kitchen_info.append(basic_kitchen)

            excel_bytes = CE.generate_sheet(basic_kitchen_info, genInfo)
            if excel_bytes is None:
                st.error("Error generating Excel sheet: No data returned.")
                return
            
            st.download_button(
                label="⬇️ Download Excel File",
                data=excel_bytes.getvalue(),
                file_name=f"{genInfo['projectNum']} Cost Sheet {genInfo['date']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.info("1. Download the Excel file\n2. Fill in the detailed specifications in Excel\n3. Let Excel calculate the values\n4. Save the file and upload below")
            
        except Exception as e:
            st.error(f"Error generating Excel: {str(e)}")

    # Upload section
    st.markdown("### Step 2: Upload Processed Excel")
    uploaded_file = st.file_uploader("Upload Excel file after calculations", type=['xlsx'])
    
    if uploaded_file is not None:
        try:
            st.write("Loading Excel file...")
            excel_data = BytesIO(uploaded_file.getvalue())
            
            # Load template workbook for dropdowns
            template_path = TEMPLATES['EXCEL']
            template_wb = openpyxl.load_workbook(template_path)
            
            # Load workbooks for processing
            wb_data_with_formulas = openpyxl.load_workbook(excel_data)
            excel_data.seek(0)
            wb_data = openpyxl.load_workbook(excel_data, data_only=True)
            
            # Extract data and update totals as before
            extracted_data = extract_data_from_excel(wb_data)
            
            # Update JOB TOTAL and general info
            if 'JOB TOTAL' in wb_data_with_formulas.sheetnames:
                job_total_sheet = wb_data_with_formulas['JOB TOTAL']
                
                # Write totals
                job_total_sheet['S16'] = extracted_data['total_costs']
                job_total_sheet['T16'] = extracted_data['total_job_price']
                
                # Write general info
                job_total_sheet['C3'] = genInfo.get('projectNum', '')
                job_total_sheet['C5'] = genInfo.get('customer', '')
                job_total_sheet['C7'] = genInfo.get('combined_initials', '')
                job_total_sheet['G3'] = genInfo.get('projectName', '')
                job_total_sheet['G5'] = genInfo.get('location', '')
                job_total_sheet['G7'] = genInfo.get('date', '')
            
            # Copy dropdowns from template to modified workbook
            copy_dropdowns(template_wb, wb_data_with_formulas)
            
            # Save the modified workbook
            modified_excel = BytesIO()
            wb_data_with_formulas.save(modified_excel)
            modified_excel.seek(0)
            
            # Generate Word document
            word_file = CW.generate_word(extracted_data, genInfo)
            
            # Create ZIP with modified Excel file
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Add modified Excel file
                zip_file.writestr(
                    f"{genInfo['projectNum']} Cost Sheet {genInfo['date']}.xlsx",
                    modified_excel.getvalue()  # Use modified Excel file
                )
                
                # Add Word file
                zip_file.writestr(
                    f"{genInfo['projectNum']} Quotation.docx",
                    word_file.getvalue()
                )
            
            # Provide download button for ZIP file
            st.download_button(
                label="⬇️ Download ZIP (Excel + Word)",
                data=zip_buffer.getvalue(),
                file_name=f"{genInfo['projectNum']} Documents.zip",
                mime="application/zip"
            )
            
            # Optional: Display some summary
            st.write("### Summary")
            st.write(f"Total Job Price: £{extracted_data['total_job_price']:,.2f}")
            for kitchen in extracted_data['kitchens']:
                st.write(f"\n{kitchen['kitchen_name']}")
                for floor in kitchen['floors']:
                    st.write(f"  {floor['floor_name']}: {len(floor['canopies'])} canopies")
            
        except Exception as e:
            st.error(f"Error processing files: {str(e)}")

    # Use the names from general_info
  
def fill_dummy_kitchen_data():
    """Fill kitchen form with dummy data for testing"""
    # Set kitchen name
    st.session_state['kitchen_name'] = "Main Kitchen"
    
    # Set floor name
    st.session_state['floor_name'] = "Ground Floor"
    
    # Set canopy data
    st.session_state['itemNum'] = "KM123-1"
    st.session_state['model'] = "KVF"  # One of the models that needs Supply Air calc
    st.session_state['width'] = 1000
    st.session_state['length'] = 2000
    st.session_state['height'] = 600
    st.session_state['section'] = 2
    st.session_state['flowRate'] = 0.5
    
    # Calculate Supply Air for specific models
    if st.session_state['model'] in ['UVX-M', 'KVX-M', 'KVF', 'CMWF', 'UVF']:
        # Supply Air calculation
        supply_air = st.session_state['flowRate'] * 0.85  # 85% of extract rate
        st.session_state['supply_air'] = supply_air
    
    # Set lights
    st.session_state['lights'] = "LED Strip Light"
    st.session_state['light_quantity'] = 2
    
    # Set wall cladding
    st.session_state['cladding'] = True
    st.session_state['cladding_desc'] = ["Rear", "Left"]

def canopy_main():
    excel_template_path = TEMPLATES['EXCEL']
    
    if not os.path.exists(excel_template_path):
        st.error(f"Excel template not found at: {excel_template_path}")
        return
    
    st.title("Canopy Cost Sheet Generator")
    
    # Add dummy data button
    if st.button("Fill Test Kitchen Data"):
        fill_dummy_kitchen_data()
    
    # Rest of your existing code...

def extract_data_from_excel(wb):
    """Extract all canopy data from uploaded Excel file"""
    kitchen_info = []
    p182_values = {}  # Store P182 values for each sheet
    n9_values = {}    # Store N9 values for each sheet
    canopy_prices = {}  # Store canopy prices for each sheet
    
    def get_numeric_value(cell_ref, sheet=None, default=0):
        """Get numeric value from cell, handling various formats"""
        try:
            value = sheet[cell_ref].value
            if isinstance(value, str):
                value = value.replace('£', '').replace(',', '').strip()
            return math.ceil(float(value)) if value else default
        except:
            return default

    # Process each CANOPY sheet
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('CANOPY - '):
            sheet = wb[sheet_name]
            
            # Store P182 and N9 values for this sheet
            p182_values[sheet_name] = get_numeric_value('P182', sheet)
            n9_values[sheet_name] = get_numeric_value('N9', sheet)
            
            # Rest of your existing code...

    # Add to extracted data
    return {
        'kitchens': kitchen_info,
        'p182_values': p182_values,
        'n9_values': n9_values,
        'canopy_prices': canopy_prices,
        'total_job_price': math.ceil(total_job_price),
        'total_costs': math.ceil(total_costs)
    }

def parse_cladding_description(desc):
    """Extract wall selections from cladding description"""
    if not desc:
        return []
    
    walls = []
    if 'Rear' in str(desc):
        walls.append('Rear')
    if 'Left' in str(desc):
        walls.append('Left')
    if 'Right' in str(desc):
        walls.append('Right')
    return walls

def copy_dropdowns(source_wb, target_wb):
    """Copy dropdown validations from source workbook to target workbook"""
    # First, ensure Lists sheet exists and copy it
    if 'Lists' in source_wb.sheetnames:
        # Copy Lists sheet if it doesn't exist in target
        if 'Lists' not in target_wb.sheetnames:
            target_wb.create_sheet('Lists')
        
        source_lists = source_wb['Lists']
        target_lists = target_wb['Lists']
        
        # Copy all values from Lists sheet
        for row in source_lists.rows:
            for cell in row:
                if cell.value is not None:
                    target_lists[cell.coordinate] = cell.value

        # Define dropdowns configuration
        dropdowns = {
            'configuration': {
                'options': ['WALL', 'ISLAND'],
                'column': 'A',
                'target_col': 'C',
                'row_offset': 2  # C14 for first canopy (12 + 2)
            },
            'model': {
                'options': ['KVF', 'KVX-M', 'KVI', 'UVX', 'UVX-M', 'UVI', 'UVF', 'UV-C POD', 'CMWI', 'CMWF', 'CXW', 'CXW-M', 'KVV'],
                'column': 'B',
                'target_col': 'D',
                'row_offset': 2  # D14 for first canopy
            },
            'lights': {
                'options': ['LED Strip Light', 'LED Light', 'LED High Power'],
                'column': 'C',
                'target_col': 'C',
                'row_offset': 3  # C15 for first canopy
            },
            'wall_cladding': {
                'options': ['', '2M² (HFL)'],
                'column': 'D',
                'target_col': 'C',
                'row_offset': 7  # C19 for first canopy
            },
            'control_panel': {
                'options': ['CP1S', 'CP2S', 'CP3S', 'CP4S'],
                'column': 'E',
                'target_col': 'C',
                'row_offset': 13  # C25 for first canopy
            },
            'ww_pods': {
                'options': ['1000-S', '1500-S', '2000-S', '2500-S', '3000-S', '1000-D', '1500-D', '2000-D', '2500-D', '3000-D'],
                'column': 'F',
                'target_col': 'C',
                'row_offset': 14  # C26 for first canopy
            }
        }

        # Create dropdowns for each CANOPY sheet
        for sheet_name in target_wb.sheetnames:
            if sheet_name.startswith('CANOPY') and sheet_name != 'CANOPY':
                sheet = target_wb[sheet_name]
                
                # Write options to Lists sheet and create validations
                for name, config in dropdowns.items():
                    # Write options to Lists sheet
                    for i, option in enumerate(config['options'], 1):
                        target_lists[f"{config['column']}{i}"] = option
                    
                    # Create range reference
                    range_ref = f"Lists!${config['column']}$1:${config['column']}${len(config['options'])}"
                    
                    # Create validation
                    dv = DataValidation(
                        type="list",
                        formula1=range_ref,
                        allow_blank=True
                    )
                    sheet.add_data_validation(dv)
                    
                    # Apply validation to cells
                    current_row = 12  # Starting row
                    while current_row <= sheet.max_row:
                        dv.add(f"{config['target_col']}{current_row + config['row_offset']}")
                        current_row += 17  # Move to next canopy section

def save_form_data(kitchen_info, genInfo):
    """Save form data to JSON"""
    data = {
        'kitchen_info': kitchen_info,
        'genInfo': genInfo
    }
    
    # Create JSON buffer
    json_buffer = BytesIO()
    json_buffer.write(json.dumps(data, indent=2).encode())
    json_buffer.seek(0)
    
    return json_buffer

def load_form_data(json_file):
    """Load form data from JSON file"""
    try:
        data = json.loads(json_file.read().decode())
        return data.get('kitchen_info', []), data.get('genInfo', {})
    except Exception as e:
        st.error(f"Error loading JSON data: {str(e)}")
        return [], {}

# In the main UI section:
st.markdown("### Save/Load Project Data")
col1, col2 = st.columns(2)

# Save current form data
with col1:
    if st.button("Save Project Data"):
        json_buffer = save_form_data(kitchen_info, genInfo)
        st.download_button(
            label="⬇️ Download Project Data",
            data=json_buffer.getvalue(),
            file_name=f"{genInfo['projectNum']}_project_data.json",
            mime="application/json"
        )

# Load saved form data
with col2:
    uploaded_json = st.file_uploader("Load Project Data", type=['json'])
    if uploaded_json:
        loaded_kitchen_info, loaded_genInfo = load_form_data(uploaded_json)
        if loaded_kitchen_info and loaded_genInfo:
            # Auto-fill form fields
            for key, value in loaded_genInfo.items():
                if key in st.session_state:
                    st.session_state[key] = value
            
            # Set kitchen count
            st.session_state['num_kitchens'] = len(loaded_kitchen_info)
            
            # Fill kitchen data
            for i, kitchen in enumerate(loaded_kitchen_info):
                st.session_state[f'kitchen_name_{i}'] = kitchen['kitchen_name']
                st.session_state[f'floors_input_{i}'] = len(kitchen['floors'])
                
                for j, floor in enumerate(kitchen['floors']):
                    st.session_state[f'floor_name_{i}_{j}'] = floor['floor_name']
                    st.session_state[f'canopies_input_{i}_{j}'] = len(floor['canopies'])
                    
                    for k, canopy in enumerate(floor['canopies']):
                        for field, value in canopy.items():
                            st.session_state[f'{field}_{i}_{j}_{k}'] = value

# Upload section for both files
st.markdown("### Step 2: Upload Files")
col1, col2 = st.columns(2)

with col1:
    uploaded_json = st.file_uploader("Upload Project Data (Optional)", type=['json'])

with col2:
    uploaded_excel = st.file_uploader("Upload Excel file", type=['xlsx'])

if uploaded_excel:
    if uploaded_json:
        # Load JSON data first
        loaded_kitchen_info, loaded_genInfo = load_form_data(uploaded_json)
        # Update genInfo with loaded data
        genInfo.update(loaded_genInfo)
    
    # Process Excel file as before
    try:
        st.write("Loading files...")
        excel_data = BytesIO(uploaded_file.getvalue())
        
        # Load template workbook for dropdowns
        template_path = TEMPLATES['EXCEL']
        template_wb = openpyxl.load_workbook(template_path)
        
        # Load workbooks for processing
        wb_data_with_formulas = openpyxl.load_workbook(excel_data)
        excel_data.seek(0)
        wb_data = openpyxl.load_workbook(excel_data, data_only=True)
        
        # Extract data and update totals as before
        extracted_data = extract_data_from_excel(wb_data)
        
        # Update JOB TOTAL and general info
        if 'JOB TOTAL' in wb_data_with_formulas.sheetnames:
            job_total_sheet = wb_data_with_formulas['JOB TOTAL']
            
            # Write totals
            job_total_sheet['S16'] = extracted_data['total_costs']
            job_total_sheet['T16'] = extracted_data['total_job_price']
            
            # Write general info
            job_total_sheet['C3'] = genInfo.get('projectNum', '')
            job_total_sheet['C5'] = genInfo.get('customer', '')
            job_total_sheet['C7'] = genInfo.get('combined_initials', '')
            job_total_sheet['G3'] = genInfo.get('projectName', '')
            job_total_sheet['G5'] = genInfo.get('location', '')
            job_total_sheet['G7'] = genInfo.get('date', '')
        
        # Copy dropdowns from template to modified workbook
        copy_dropdowns(template_wb, wb_data_with_formulas)
        
        # Save the modified workbook
        modified_excel = BytesIO()
        wb_data_with_formulas.save(modified_excel)
        modified_excel.seek(0)
        
        # Generate Word document
        word_file = CW.generate_word(extracted_data, genInfo)
        
        # Create ZIP with modified Excel file
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add modified Excel file
            zip_file.writestr(
                f"{genInfo['projectNum']} Cost Sheet {genInfo['date']}.xlsx",
                modified_excel.getvalue()  # Use modified Excel file
            )
            
            # Add Word file
            zip_file.writestr(
                f"{genInfo['projectNum']} Quotation.docx",
                word_file.getvalue()
            )
        
        # Provide download button for ZIP file
        st.download_button(
            label="⬇️ Download ZIP (Excel + Word)",
            data=zip_buffer.getvalue(),
            file_name=f"{genInfo['projectNum']} Documents.zip",
            mime="application/zip"
        )
        
        # Optional: Display some summary
        st.write("### Summary")
        st.write(f"Total Job Price: £{extracted_data['total_job_price']:,.2f}")
        for kitchen in extracted_data['kitchens']:
            st.write(f"\n{kitchen['kitchen_name']}")
            for floor in kitchen['floors']:
                st.write(f"  {floor['floor_name']}: {len(floor['canopies'])} canopies")
                
    except Exception as e:
        st.error(f"Error processing files: {str(e)}")
  