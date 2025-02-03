import openpyxl
import math
import streamlit as st
import os
import pandas as pd
import formulas
from os import path
from xlcalculator import ModelCompiler
from xlcalculator import Model
from xlcalculator import Evaluator
import xlwings as xw
from io import BytesIO

def extract_canopy_prices(excel_path):
    """
    Extract price from cell P12 in the Excel file.
    """
    try:
        if not os.path.exists(excel_path):
            st.error(f"Excel file not found: {excel_path}")
            return []
        
        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        # Get the value from P12
        total_value = sheet['P12'].value
        
        if total_value is not None:
            # Convert to float if it's a string
            if isinstance(total_value, str):
                try:
                    total_value = float(total_value.replace(',', ''))
                except ValueError:
                    st.warning(f"Could not convert value '{total_value}' to number")
                    return []
            
            # Round up to nearest integer
            if isinstance(total_value, (int, float)):
                rounded_value = math.ceil(total_value)
                return [{
                    'canopy_number': 1,
                    'total_price': rounded_value
                }]
        
        wb.close()
        return []
        
    except Exception as e:
        st.error(f"Error extracting price: {str(e)}")
        return []

def convert_formulas_to_values(excel_path):
    """
    Opens Excel file and runs the ConvertFormulasToValues macro
    """
    try:
        # Create Excel application object
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Run in background
        
        # Open workbook
        wb = excel.Workbooks.Open(excel_path)
        
        # Run the macro
        excel.Run("ConvertFormulasToValues")
        
        # Save and close
        wb.Save()
        wb.Close()
        excel.Quit()
        
    except Exception as e:
        print(f"Error converting formulas: {str(e)}")
        if 'excel' in locals():
            excel.Quit()

def get_excel_calculated_values(excel_path):
    """
    Reads Excel file and saves as CSV for inspection.
    """
    try:
        # Read Excel file
        df = pd.read_excel(excel_path, sheet_name='CANOPY', header=None)
        
        # Save as CSV in the same directory
        csv_path = excel_path.replace('.xlsx', '.csv')
        df.to_csv(csv_path, index=False)
        
        # Print file location and first few rows
        print(f"\nCSV saved at: {csv_path}")
        print("\nFirst 20 rows of CSV:")
        print(df.head(20).to_string())
        
        # Print specific rows we're interested in
        print("\nSpecific Rows:")
        print("\nRow 183 (Delivery):")
        print(df.iloc[182].to_string())
        print("\nRow 189 (Installation):")
        print(df.iloc[188].to_string())
        print("\nRow 200 (Total):")
        print(df.iloc[199].to_string())
        
        # Print canopy rows
        print("\nCanopy Rows (every 17 rows):")
        for row in range(10, len(df), 17):
            print(f"\nRow {row+1}:")
            print(df.iloc[row].to_string())
        
        return csv_path
        
    except Exception as e:
        print(f"Error reading Excel values: {str(e)}")
        print(f"At path: {excel_path}")
        return None

def run_excel_script(excel_path, macro_name=None):
    """
    Reads N12 calculated value and pastes to P12
    """
    try:
        st.write("Processing Excel file...")
        
        # Read Excel file with pandas to get calculated values
        df = pd.read_excel(excel_path, sheet_name='CANOPY', header=None)
        
        # N12 is at row 11, column 13 (0-based indexing)
        n12_value = df.iloc[11, 13]  # This will get the calculated value
        st.write(f"Read calculated value from N12: {n12_value}")
        
        if not pd.isna(n12_value):  # Check if value exists
            # Load workbook to write the value
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb['CANOPY']
            
            # Write the calculated value to P12
            sheet['P12'].value = float(n12_value)
            st.write(f"Wrote value to P12: {n12_value}")
            
            # Save workbook
            wb.save(excel_path)
            wb.close()
            
            return float(n12_value)
            
        else:
            st.warning("No value found in N12")
            return None
        
    except Exception as e:
        st.error(f"Error processing Excel file: {str(e)}")
        return None

def convert_to_macro_enabled(excel_bytes):
    """
    Converts Excel bytes to a macro-enabled workbook
    """
    try:
        st.write("Converting to macro-enabled workbook...")
        
        # Create a temporary file to write the bytes
        temp_path = "temp_original.xlsx"
        with open(temp_path, "wb") as f:
            f.write(excel_bytes.getvalue())
        
        # Load the workbook
        wb = openpyxl.load_workbook(temp_path)
        
        # Save as macro-enabled workbook
        macro_path = "temp_macro.xlsm"
        wb.save(macro_path)
        wb.close()
        
        # Read the macro-enabled file back
        with open(macro_path, "rb") as f:
            macro_bytes = BytesIO(f.read())
        
        # Clean up temp files
        os.remove(temp_path)
        os.remove(macro_path)
        
        st.write("Conversion complete")
        return macro_bytes
        
    except Exception as e:
        st.error(f"Error converting to macro-enabled: {str(e)}")
        return None

def read_excel_value(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        value = ws['N12'].value
        wb.close()
        return value
    except Exception as e:
        st.error(f"Error reading Excel: {str(e)}")
        return None